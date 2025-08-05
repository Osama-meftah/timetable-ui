from ortools.sat.python import cp_model
from collections import defaultdict
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl import Workbook
import random
from .models import Department, Program, Hall, Level, Group, Subject, Teacher,Period,Today,TeacherTime, Distribution, Table
from tempfile import NamedTemporaryFile
import hashlib

class TimeTableScheduler:
    def __init__(self,semester_filter=None):
        self.semester_filter = semester_filter
        self.model                  = cp_model.CpModel() 
        self.solver                 = cp_model.CpSolver()

        self.available_times_df     =list(Period.objects.all())
        # self.periods                = Period.objects.all()
        self.professors              = {p.id: p for p in Teacher.objects.filter(teacher_status="active")}
        self.courses                 = {c.id: c for c in Subject.objects.all()}
        self.rooms = list(Hall.objects.filter(hall_status="available"))
        self.days = list(Today.objects.all())
        self.timesProfessor =list(TeacherTime.objects.all())
        if self.semester_filter:
            self.teatchingGroups = list(Distribution.objects.filter(fk_subject__term=self.semester_filter))
        else:
            self.teatchingGroups = list(Distribution.objects.all())
        # self.teatchingGroups =list(Distribution.objects.all())
        self.programs = {p.id: p for p in Program.objects.all()}
        self.levels = {l.id: l for l in Level.objects.all()}
        self.groups = {g.id: g for g in Group.objects.all()}

        self.lecture_times          ={}   
        self.professoresdata        = []
        self.generated_schedule = []
        self.available_times        = {}
        self.temp_file = None
        self.log=[]
        self.conflicts=[]
        self.available_unscheduled_slots=[]
        self.random_enabled =True
        self.penalties = []
    def add_data(self): 
        for row in self.available_times_df:
            time_from=row.period_from
            time_to=row.period_to
            time = f"{time_from}-{time_to}"
            self.available_times[row.pk]=time
        for tg in self.teatchingGroups:
            professor_id = tg.fk_teacher.id
            course_id = tg.fk_subject.id
            group_id = tg.fk_group.pk

            professor = self.professors[professor_id]
            course = self.courses[course_id]
            group = self.groups[group_id]
            level = self.levels[group.fk_level.pk]
            dept = self.programs[level.fk_program.pk]

            professor_times = [t for t in self.timesProfessor if t.fk_teacher.id == professor_id]

            availability = defaultdict(list)
            for t in professor_times:
                availability[t.fk_today.pk].append(t.fk_period.pk)

            professordata = {
                "available": dict(availability),
                "name": professor.teacher_name.strip(),
                "ProfessorId": professor_id,
                "email":professor.user.email
            }

            self.professoresdata.append(professordata)

            self.lecture_times[tg.id] = {
                'course': course.subject_name.strip(),
                'teacher': professordata,
                'level': level.get_level_name_display().strip(),
                'dept': dept.program_name.strip(),
                'group': group.group_name.strip(),
                'std_count': group.number_students
            }
    def define_variables(self):
        self.schedule_vars = {}
        for course_id, course_info in self.lecture_times.items():
            for day in self.days: 
                for time_index in self.available_times_df:  
                    for room in self.rooms:  
                        var_name = f'course_{course_id}_day_{day.pk}_time_{time_index.pk}_room_{room.hall_name}'
                        self.schedule_vars[(course_id, day.pk, time_index.pk, room.hall_name)] = self.model.NewBoolVar(var_name)

    def add_constraints(self):
        self.add_lecture_once_soft_constraint()
        self.add_room_time_constraints()
        self.add_teacher_constraints()
        self.add_dept_level_group_time_constraints()
        self.add_teacher_constraints_availability()
        self.model.Minimize(sum(self.penalties))

    # قيد عدم تكرار المحاضرة في الاسبوع
    def add_lecture_once_soft_constraint(self):
        """القيد 1 (مرن): كل محاضرة يجب أن تحدث مرة واحدة. إذا لم يحدث، ندفع عقوبة."""
        for course_id, course_info in self.lecture_times.items():
            all_possible_slots = [self.schedule_vars[(course_id, day.pk, time_index.pk, room.hall_name)]
                                  for day in self.days
                                  for time_index in self.available_times_df
                                  for room in self.rooms]
            is_scheduled = self.model.NewBoolVar(f'is_scheduled_{course_id}')
            self.model.Add(sum(all_possible_slots) == is_scheduled)
            self.penalties.append(is_scheduled.Not())
    
    # قيد توزيع القاعات حسب السعة وبحيث لا تكرر المحاضره في اكثر من قاعة 
    def add_room_time_constraints(self):
        for day in self.days:
            for time_index in self.available_times_df:
                for row in self.rooms:
                    capacity=row.capacity_hall
                    room_name=row.hall_name
                    lectures_in_room=[]
                    for course_id, course_info in self.lecture_times.items():
                        student_count = course_info['std_count']
                        if student_count > capacity:
                            self.model.Add(self.schedule_vars[(course_id, day.pk, time_index.pk, room_name)]==0)

                        lectures_in_room.append(self.schedule_vars[(course_id, day.pk, time_index.pk, room_name)])
                    self.model.AddAtMostOne(lectures_in_room)

    # قيد عدم تكرار المدرس في نفس الوقت
    def add_teacher_constraints(self):
        for teacher_name in self.professoresdata:
            teacher_name = teacher_name['name']
            for day in self.days:
                for time_index in self.available_times_df:
                    lectures = []
                    for course_id, course_info in self.lecture_times.items():
                        if course_info['teacher']["name"] == teacher_name:
                            # print("========================= ",course_info['teacher']["name"])
                            for room in self.rooms:
                                lectures.append(self.schedule_vars[(course_id, day.pk, time_index.pk, room.hall_name)])
                    self.model.Add(sum(lectures) <= 1) 


    def add_teacher_constraints_availability(self):
        for course_id, course_info in self.lecture_times.items():
           available_times = course_info['teacher']["available"]
           for day in self.days:
              for time_index in self.available_times_df:
                  is_available = day.pk in available_times and time_index.pk in available_times.get(day.pk, [])
                  if not is_available:
                      for room in self.rooms:
                         self.model.Add(self.schedule_vars[(course_id, day.pk, time_index.pk, room.hall_name)] == 0)

 # إضافة قيود لمنع تكرار نفس القسم والمستوى والمجموعه في نفس الوقت
    def add_dept_level_group_time_constraints(self):
        # تجميع المواد حسب القسم والمستوى
        dept_level_courses = defaultdict(list)
        for course_id, course_info in self.lecture_times.items():
            key = (course_info['dept'], course_info['level'], course_info['group'])
            dept_level_courses[key].append(course_id)

        for (dept, level,group), courses in dept_level_courses.items():
            for day in self.days:
                for time_index in self.available_times_df:
                    time_slot_vars = []
                    for course_id in courses:
                        for room in self.rooms:
                            time_slot_vars.append(self.schedule_vars[(course_id, day.pk, time_index.pk, room.hall_name)])
                    # لا يمكن جدولة أكثر من مادة لنفس القسم والمستوى في نفس الوقت
                    if time_slot_vars:
                        self.model.Add(sum(time_slot_vars) <= 1)
    
    @staticmethod
    def set_cell_border(cell):
        """
        تعيين الحدود لخلية معينة في Excel.
        """
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.border = thin_border

    # save schedule in excel file
    def convert_unscheduled_to_conflicts(self, unscheduled_lectures):

        professor_courses=defaultdict(list)
        for course_id, course_info in self.lecture_times.items():
            professor_name = course_info['teacher']['name']
            professor_courses[professor_name].append(course_info['course'])
        
        if not unscheduled_lectures:
            return []

        converted_conflicts = []
        for info in unscheduled_lectures:
            teacher_name = info['teacher']['name']
            email=info['teacher']['email']
            course_name = info['course']
            group_info = f"{info['dept']}-{info['level']}-{info['group']}"
            std_count = info['std_count']
            available = info['teacher']['available']
            num_available_course=f"({len(professor_courses[teacher_name])}) مقرر  -  ({len(available)}) وقت متاح"

            # تحويل الأوقات إلى نص مفهوم
            readable_times = []
            for day_id, time_indices in available.items():
                # الحصول على اسم اليوم من self.days
                matching_day = next((d for d in self.days if d.pk == day_id), None)
                if matching_day:
                    day_name = matching_day.get_day_name_display()
                    for time_idx in time_indices:
                        if time_idx < len(self.available_times):
                            time_str = self.available_times[time_idx]
                            readable_times.append(f"{day_name} {time_str}")

            detail_text = {
                "teacher":teacher_name,
                'email':email,
                "course":course_name,
                "group":group_info,
                "num_available_course":num_available_course,
                "available":f"{', '.join(readable_times) if readable_times else 'لا توجد'}",
                "std_count":std_count
            }

            converted_conflicts.append({
                'conflicts_type': 'محاضرة غير مجدولة',
                'conflicts_detail': detail_text
            })

            self.conflicts=converted_conflicts

    def save_to_excel(self, schedule):
        wb = Workbook()
        ws = wb.active
        ws.title = "جدول المقررات الدراسية"
        rooms = [r.hall_name for r in self.rooms]
        headers = ['Day', 'Time'] + rooms
        ws.append(headers)
        ws.row_dimensions[1].height = 80
        ws.column_dimensions['A'].width=15
        for cell in ws[1]:
            cell.font = Font(size=20, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self.set_cell_border(cell)

        schedule_by_day = defaultdict(list)
        for row in schedule:
            schedule_by_day[row['day']].append(row)

        # 🟨 تخصيص لون عشوائي لكل برنامج (قسم)
        pastel_colors = [
        "E6F7FF",  # Light Blue
        "E8F5E9",  # Light Green
        "FFF3E0",  # Light Orange
        "F3E5F5",  # Light Purple
        "FFFDE7",  # Light Yellow
        "FCE4EC",  # Light Pink
        "E0F7FA",  # Light Cyan
        "F9FBE7",  # Light Lime
        "F5F5F5",  # Light Gray
        "E1F5FE",  # Pale Blue
        "FFE0B2",  # Peach
        "E1BEE7",  # Lavender
        "D1C4E9",  # Soft Purple
        "C8E6C9",  # Mint Green
        "DCEDC8",  # Pastel Green
        "FFECB3",  # Light Gold
        "FFCDD2",  # Blush Pink
        "D7CCC8",  # Soft Brown
        "F8BBD0",  # Baby Pink
        "D0F0FD",  # Sky Blue
        "E6EE9C",  # Lime Pastel
        "B2EBF2",  # Aqua Pastel
        "B3E5FC",  # Cool Blue
        "FFCCBC",  # Apricot
        "E0F2F1",  # Teal Tint
        "EDE7F6",  # Violet Gray
        "FFF9C4",  # Cream Yellow
        "F3EDE3",  # Light Beige
        "F0F4C3",  # Lemon Pastel
        "FFE082",  # Banana Yellow
        "C5CAE9",  # Steel Lilac
        "F0F0F0",  # Ultra Light Gray
        "E3F2FD",  # Icy Blue
        "F1F8E9",  # Avocado Light
        "FFEBEE",  # Rose White
        "FFF8E1",  # Light Custard
        "F9FBE7",  # Lemon Ice
        "E3FDFD",  # Cool Mist
        "FFFAF0",  # Floral White
        "F0FFFF",  # Azure White
    ]

        colors_by_program = {}
        def get_color_for_program(program_name):
            if program_name in colors_by_program:
                return colors_by_program[program_name]
            # استخدم hash ثابت من اسم القسم
            hash_value = int(hashlib.sha256(program_name.encode('utf-8')).hexdigest(), 16)
            color_index = hash_value % len(pastel_colors)
            color = pastel_colors[color_index]
            colors_by_program[program_name] = color
            return color

        row_num = 2
        for day, rows in schedule_by_day.items():
            start_merge_row = row_num
            end_merge_row = row_num + len(self.available_times) - 1
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=1)
            day_cell = ws.cell(row=start_merge_row, column=1, value=day)
            day_cell.font = Font(size=20, bold=True)
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.set_cell_border(day_cell)

            for slot_id, slot_value in self.available_times.items():
                ws[f'B{row_num}'] = slot_value
                for col in range(1, len(self.rooms) + 3):
                    
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.cell(row=row_num, column=col).font = Font(name="Arial", size=20)
                    if col!=1:
                        col_letter = ws.cell(row=1, column=col).column_letter
                        ws.column_dimensions[col_letter].width = 30

                added = False
                for row in rows:
                    if row['time'] == slot_value:
                        for j, room in enumerate(self.rooms, start=2):
                            if row['room'] == room.hall_name:
                                program = row['dept']
                                fill_color = get_color_for_program(program)
                                fill_color = colors_by_program[program]

                                cell_value = f"{row['course']}\n{row['dept']}_{row['level']}_{row['group']}\n{row['teatcher']}"
                                cell = ws.cell(row=row_num, column=j + 1, value=cell_value)
                                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                added = True
                if not added:
                    for j in range(2, len(self.rooms) + 2):
                        ws.cell(row=row_num, column=j + 1, value="")

                first_row = False
                for col in range(1, len(self.rooms) + 3):
                    cell = ws.cell(row=row_num, column=col)
                    self.set_cell_border(cell)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(name="Arial", size=20)
                    if col !=1:
                        col_letter = ws.cell(row=1, column=col).column_letter
                        ws.column_dimensions[col_letter].width = 30
                row_num += 1

        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.seek(0)
        return temp_file
    
    def process_solution(self):
            scheduled = []
            unscheduled = []
            scheduled_ids = set()
            scheduled_slots = set() 

            # Existing logic for scheduled lectures
            for day in self.days:
                for time in self.available_times_df:
                    for room in self.rooms:
                        capacity = room.capacity_hall
                        room_name = room.hall_name
                        for course_id, course_info in self.lecture_times.items():
                            if self.solver.Value(self.schedule_vars[(course_id, day.pk, time.pk, room_name)]) == 1:
                                id = time.pk
                                scheduled.append({
                                        "course_id": course_id,
                                        "day": day.get_day_name_display(),
                                        "time": self.available_times[id],
                                        "room": room_name,
                                        "room_id": room.pk,
                                        "capacity_room": capacity,
                                        "course": course_info['course'],
                                        "teatcher": course_info['teacher']['name'],
                                        "available": course_info['teacher']['available'],
                                        "day_id": day.pk,
                                        "time_id": id,
                                        "group": course_info['group'],
                                        "level": course_info['level'],
                                        "dept": course_info['dept'],
                                        "student_count": course_info['std_count']
                                })
                                scheduled_ids.add(course_id)
                                scheduled_slots.add((day.pk, time.pk, room.pk)) 

            # Populate unscheduled lectures
            for course_id, info in self.lecture_times.items():
                if course_id not in scheduled_ids:
                    unscheduled.append(info)
            # إنشاء مجموعة من جميع الفتحات المحتملة
            all_possible_slots = set()
            for day in self.days:
                for time_obj in self.available_times_df:
                    for room in self.rooms:
                        all_possible_slots.add((day.pk, time_obj.pk, room.pk))

            # الفرق بين جميع الفتحات الممكنة والفتحات المجدولة هو الفتحات غير المجدولة
            # الفتحات غير المجدولة هي الأوقات/الغرف التي كانت متاحة لكن لم يُسند إليها أي محاضرة
            for day_pk, t_id, room_pk in all_possible_slots:
                if (day_pk, t_id, room_pk) not in scheduled_slots:
                    # ابحث عن تفاصيل اليوم والوقت والغرفة لعرضها
                    day_name = next((d.get_day_name_display() for d in self.days if d.pk == day_pk), "غير معروف")
                    time_name = self.available_times.get(t_id, "غير معروف")
                    room_name = next((r.hall_name for r in self.rooms if r.pk == room_pk), "غير معروف")
                    
                    self.available_unscheduled_slots.append({
                        "day": day_name,
                        "time": time_name,
                        "room": room_name
                    })
                
            return scheduled, unscheduled # Return the new list
    def solve(self):
        
        if self.random_enabled:
            self.solver.parameters.random_seed=random.randint(1,10000)
        else:
            self.solver.parameters.random_seed=0

        self.solver.parameters.enumerate_all_solutions=False
        status = self.solver.Solve(self.model)
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            scheduled_lectures, unscheduled_lectures = self.process_solution()
            
            if not unscheduled_lectures:
                self.log.append("\n✅ تم إيجاد حل مثالي! تم جدولة جميع المحاضرات بنجاح.")
                self.temp_file=self.save_to_excel(scheduled_lectures)
                self.generated_schedule=scheduled_lectures

            else:
                self.log.append(f"\n🟡 تم إيجاد أفضل حل ممكن. تم جدولة {len(scheduled_lectures)} محاضرة بنجاح.")
                self.convert_unscheduled_to_conflicts(unscheduled_lectures)

        else:
            self.log.append(f"❌ فشل الحل بشكل غير متوقع حتى مع القيود المرنة (الحالة: {self.solver.StatusName(status)})")

    def run(self):
        self.add_data()
        self.define_variables()
        self.add_constraints()
        self.solve()

