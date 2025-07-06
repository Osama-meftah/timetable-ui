import pandas as pd
from ortools.sat.python import cp_model
from collections import defaultdict
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl import Workbook
import random

class TimeTableScheduler:
    def __init__(self):
        self.model                  = cp_model.CpModel() 
        self.solver                 = cp_model.CpSolver()
        
        self.available_times        = []

        self.available_times_df     = pd.read_excel('data/period.xlsx')
        self.professors             = pd.read_excel('data/Professors.xlsx')  
        self.courses                = pd.read_excel('data/Courses.xlsx')  
        self.rooms                  = pd.read_excel('data/Rooms.xlsx')  
        self.days                   = pd.read_excel('data/Days.xlsx')
        self.timesProfessor         = pd.read_excel('data/ProDayTimes.xlsx')  
        self.teatchingGroups        =pd.read_excel("data/teaching_group.xlsx")
        self.programs               =pd.read_excel("data/programs.xlsx")
        self.levels                 =pd.read_excel("data/levels.xlsx")  
        self.groups                 =pd.read_excel("data/groups.xlsx")
        self.lecture_times          ={}   
        self.professoresdata        = []
    
    def add_data(self): 
        for _,row in self.available_times_df.iterrows():
            time_from=row['period_from']
            time_to=row['period_to']
            time = f"{time_from}-{time_to}"
            self.available_times.append(time)

        for _, row in self.teatchingGroups.iterrows():
            professor_id = row['ProfessorID']
            course_id = row['CourseID']
            group_id = row['GroupID']
            
            professor_name = self.professors.loc[self.professors['ProfessorID'] == professor_id, 'ProfessorName'].iloc[0]
            course_name = self.courses.loc[self.courses['CourseID'] == course_id, 'Subject'].iloc[0] 
            professor_times = self.timesProfessor.loc[self.timesProfessor['ProfessorID'] == professor_id].values
            student_count=self.groups.loc[self.groups['id']==group_id, 'std_count'].iloc[0] 
            levelsId=self.groups.loc[self.groups['id']==group_id, 'level_id'].iloc[0] 
            level_name = self.levels.loc[self.levels['id'] == levelsId, 'name'].iloc[0] 
            deptId=self.levels.loc[self.levels['id'] == levelsId, 'program_id'].iloc[0]
            dept_name = self.programs.loc[self.programs['id'] == deptId, 'name'].iloc[0]
            group_name=self.groups.loc[self.groups['id'] == group_id, 'name'].iloc[0]

            professordata = {
                "available":{
                    pro_time[2]:[i - 3 for i, val in enumerate(pro_time[3:], start=3) if val > 0]
                    for pro_time in professor_times
                    }
                }
            professordata["name"] = professor_name.strip()
            professordata["ProfessorId"] = professor_id            
            self.professoresdata.append(professordata)
            self.lecture_times[row['teatch_id']] = {
                'course':course_name.strip(),
                'teacher': professordata,
                'level': level_name,
                'dept': dept_name,
                'group': group_name,
                'std_count': student_count
            }
    def define_variables(self):
        self.schedule_vars = {}
        for course_id, course_info in self.lecture_times.items():
            for day in self.days['Day Number']: 
                for time_index in range(len(self.available_times)):  
                    for room in self.rooms['room_name']:  
                        var_name = f'course_{course_id}_day_{day}_time_{time_index}_room_{room}'
                        self.schedule_vars[(course_id, day, time_index, room)] = self.model.NewBoolVar(var_name)

    def add_constraints(self):
        self.add_courses_constraints()
        self.add_room_time_constraints()
        self.add_teacher_constraints()
        self.add_teacher_constraints_availability()
        self.add_dept_level_group_time_constraints()

    # قيد عدم تكرار المحاضرة في الاسبوع
    def add_courses_constraints(self):
        for course_id, course_info in self.lecture_times.items():
            course = []
            for day in self.days['Day Number']:
                for time_index in range(len(self.available_times)):  
                    times=[]
                    for room in self.rooms['room_name']:
                        course.append(self.schedule_vars[(course_id, day, time_index, room)])  
                        times.append(self.schedule_vars[(course_id, day, time_index, room)])
                    # self.model.Add(sum(times)<=1)  
            self.model.Add(sum(course) == 1)
    
    # قيد توزيع القاعات حسب السعة وبحيث لا تكرر المحاضره في اكثر من قاعة 
    def add_room_time_constraints(self):
        for day in self.days['Day Number']:
            for time_index in range(len(self.available_times)):
                for _,row in self.rooms.iterrows():
                    capacity=row['capacity']
                    room_name=row['room_name']
                    lectures_in_room=[]
                    for course_id, course_info in self.lecture_times.items():
                        student_count = course_info['std_count']
                        if student_count > capacity:
                            self.model.Add(self.schedule_vars[(course_id, day, time_index, room_name)]==0)

                        lectures_in_room.append(self.schedule_vars[(course_id, day, time_index, room_name)])
                    self.model.AddAtMostOne(lectures_in_room)

    # قيد عدم تكرار المدرس في نفس الوقت
    def add_teacher_constraints(self):
        for teacher_name in self.professoresdata:
            teacher_name = teacher_name['name']
            for day in self.days['Day Number']:
                for time_index in range(len(self.available_times)):
                    lectures = []
                    for course_id, course_info in self.lecture_times.items():
                        if course_info['teacher']["name"] == teacher_name:
                            # print("========================= ",course_info['teacher']["name"])
                            for room in self.rooms['room_name']:
                                lectures.append(self.schedule_vars[(course_id, day, time_index, room)])
                    self.model.Add(sum(lectures) <= 1) 

    # قيد تعيين المدرس في الاوقات المتاحة لة
    def add_teacher_constraints_availability(self):
        for course_id, course_info in self.lecture_times.items():
           available_times = course_info['teacher']["available"]
            # تكرار لكل مادة
           assigned_times = []  # قائمة لتخزين المتغيرات التي تخص تعيين الدكتور في هذه المادة
           for day in self.days['Day Number']:
              for time_index in range(len(self.available_times)):
                  for room in self.rooms['room_name']:
                     var = self.schedule_vars.get((course_id, day, time_index, room))                 
                     if day in available_times and time_index in available_times[day]:
                        assigned_times.append(var)
           if assigned_times:
               self.model.Add(sum(assigned_times) >= 1)  # يجب أن يتم تعيينه على الأقل مرة واحدة للمادة  

 # إضافة قيود لمنع تكرار نفس القسم والمستوى والمجموعه في نفس الوقت
    def add_dept_level_group_time_constraints(self):
        # تجميع المواد حسب القسم والمستوى
        dept_level_courses = defaultdict(list)
        for course_id, course_info in self.lecture_times.items():
            key = (course_info['dept'], course_info['level'], course_info['group'])
            dept_level_courses[key].append(course_id)

        for (dept, level,group), courses in dept_level_courses.items():
            for day in self.days['Day Number']:
                for time_index in range(len(self.available_times)):
                    time_slot_vars = []
                    for course_id in courses:
                        for room in self.rooms['room_name']:
                            time_slot_vars.append(self.schedule_vars[(course_id, day, time_index, room)])
                    # لا يمكن جدولة أكثر من مادة لنفس القسم والمستوى في نفس الوقت
                    if time_slot_vars:
                        self.model.Add(sum(time_slot_vars) <= 1)
    
    # check conflicts after create schedule
    def check_conflicts(self, schedule):
        conflicts = []
        available_times_str = []
        for entry in schedule:
            day = entry["day"]
            day_id=entry["day_id"]
            time = entry["time"]
            doctor_name = entry["teatcher"]
            course = entry["course"]
            room=entry['room']
            std_count=entry['student_count']
            dept=entry['dept']
            level=entry['level']
            capacity=entry['capacity_room']
            available=entry['available']

            if day_id not in available:
                for day_a, times in available.items():
                    day_name = self.days['Day Name'].loc[int(day_a)]
                    for time_idx in times:
                        time_slot = f"{day_name} {self.available_times[time_idx]}"
                        if time_slot not in available_times_str:
                            available_times_str.append(time_slot)
                conflicts.append({
                    'نوع التعارض': 'تم تعيين دكتور في غير موعدة',
                    'تفاصيل التعارض':f"اسم الدكتور: {doctor_name} , اليوم:{day} , الوقت: {time} الأوقات المتاحة: {available_times_str}"})

            if(std_count>capacity):
                conflicts.append({
                    'نوع التعارض': 'لم يتم تعين قاعة',
                    'تفاصيل التعارض':f'قسم: {dept} , مستوى:{level} , عدد الطلاب: {std_count} اكبر من {capacity} في قاعة :{room}'
                })
        return conflicts
    
    #save conflicts in file excel 
    def write_conflicts_to_excel(self, conflicts):
        if conflicts:
            # Create a new Excel writer object
            conflict_writer = pd.ExcelWriter('conflicts.xlsx', engine='openpyxl')
            
            # Convert conflicts to DataFrame
            conflict_df = pd.DataFrame(conflicts)
            
            # Write to Excel with Arabic column names
            conflict_df.to_excel(conflict_writer, sheet_name='تعارضات الدكاترة', index=False)
            
            # Get the worksheet
            worksheet = conflict_writer.sheets['تعارضات الدكاترة']
            
            # Format the worksheet
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            # Save the conflicts Excel file
            conflict_writer.close()


    # check conflicts be create schedule
    def check_initial_conflicts(self):
        total_cells=len(self.available_times)*len(self.rooms)*len(self.days)
        total_lectures=len(self.lecture_times)
        print(total_cells,total_lectures)
        conflicts = []
        # Group courses by professor
        professor_courses = {}
        # conflicts_room=[]
        max_capacity = self.rooms['capacity'].max()
        dept_level_courses = defaultdict(list)
        for course_id, course_info in self.lecture_times.items():
            key = (course_info['dept'], course_info['level'])
            std_count=course_info['std_count']
            dept_level_courses[key].append(std_count)
        
        for (dept, level), stdcount in dept_level_courses.items():
            if stdcount[0]>max_capacity:
                 conflicts.append({
                    'نوع التعارض': 'لم يتم تعين قاعة',
                    'تفاصيل التعارض':f'قسم: {dept} , مستوى:{level} , عدد الطلاب: {stdcount[0]} اكبر من {max_capacity} ',
                })
        
        for course_id, course_info in self.lecture_times.items():
            professor_name = course_info['teacher']['name']
            if professor_name not in professor_courses:
                professor_courses[professor_name] = []
            professor_courses[professor_name].append({
                'course': course_info['course'],
                'available_times': course_info['teacher']['available']
            })
        
        # Check conflicts for each professor
        for professor_name, courses in professor_courses.items():
            # Count total available time slots
            total_available_slots = 0
            available_times_str = []
            
            # Get all available time slots for this professor
            for course in courses:
                for day, times in course['available_times'].items():
                    day_name = self.days['Day Name'].loc[int(day)]
                    for time_idx in times:
                        time_slot = f"{day_name} {self.available_times[time_idx]}"
                        if time_slot not in available_times_str:
                            available_times_str.append(time_slot)
                            total_available_slots += 1
            
            # Count number of courses
            num_courses = len(courses)
            # If professor has more courses than available time slots
            if num_courses > total_available_slots and total_available_slots!=0 :
                conflicts.append({
                    'نوع التعارض': 'عدد المواد أكثر من الأوقات المتاحة',
                    'تفاصيل التعارض':f'اسم الدكتور:{professor_name} ,عدد المواد:{str(num_courses)},عدد الأوقات المتاحة:{str(total_available_slots)}'
                })
            
            # If professor has no available times for some courses
            if total_available_slots == 0 and num_courses > 0:
                conflicts.append({
                    'نوع التعارض': 'لا يوجد أوقات متاحة',
                    'تفاصيل التعارض':f'اسم الدكتور:{professor_name} ,عدد المواد:{str(num_courses)},عدد الأوقات المتاحة:0'})
                
            if total_lectures>total_cells:
                  conflicts.append({
                    'نوع التعارض': 'لا يوجد قاعات كافيه ',
                    'تفاصيل التعارض':f' عدد المحاضرات كامله : {total_lectures} عدد الخلايا المتاحه في الجدول لاستيعاب المحاضرات : {total_cells}'})
                  break
            
        return conflicts
    

    @staticmethod
    def set_cell_border(cell):
        """
        تعيين الحدود لخلية معينة في Excel.
        """
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.border = thin_border

    # save schedule in excel file
    def save_to_excel(self, schedule):
        wb = Workbook()
        ws = wb.active
        ws.title = "جدول المقررات الدراسية"
        headers = ['Day', 'Time'] + self.rooms['room_name'].astype(str).tolist() 
        ws.append(headers)
        ws.row_dimensions[1].height = 80
        for cell in ws[1]:
            cell.font = Font(size=20,bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  
            cell.alignment = Alignment(horizontal="center", vertical="center",)        
        
        schedule_by_day = defaultdict(list)
        for row in schedule:
            schedule_by_day[row['day']].append(row)
        # time_slots = ['08:00-10:00', '10:00-12:00', '12:00-02:00']

        row_num = 2  
        for day, rows in schedule_by_day.items():
            first_row = True  

            for slot in self.available_times:
                if first_row:
                    ws[f'A{row_num}'] = day  
                ws[f'B{row_num}'] = slot  

                added = False 
                for row in rows:
                    if row['time'] == slot:
                        for j, room in enumerate(self.rooms['room_name'], start=2): 
                            if row['room'] == room:  
                                # dept = row['Department'].split()
                                cell_value = f"{row['course']}\n{row['dept']}_{row['level']}_{row['group']}\n{row['teatcher']}"
                                ws.cell(row=row_num, column=j + 1, value=cell_value)
                                added = True
                if not added:
                    for j in range(2, len(self.rooms) + 2):  
                        ws.cell(row=row_num, column=j + 1, value="")

                first_row = False  
                for col in range(1, len(self.rooms) + 3):  
                    cell = ws.cell(row=row_num, column=col)
                    self.set_cell_border(cell)
                    
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(name="Arial", size=20)
                    col_letter = ws.cell(row=1, column=col).column_letter  
                    ws.column_dimensions[col_letter].width = 30  
                row_num += 1
        wb.save('output/timetable.xlsx')
        print("📂 تم حفظ الجدول في 'timetable.xlsx'")    

    def solve(self):
        self.solver.parameters.random_seed=random.randint(1,10000)
        self.solver.parameters.enumerate_all_solutions=False
       
        status = self.solver.Solve(self.model)
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            print("\n✅ تم إيجاد حل!\n")
            schedule = []
            for day in self.days['Day Number']:
                for time_index, time in enumerate(self.available_times):
                    for _,row in self.rooms.iterrows():
                        capacity=row['capacity']
                        room=row['room_name']
                        for course_id, course_info in self.lecture_times.items():
                            if self.solver.Value(self.schedule_vars[(course_id, day, time_index, room)]) == 1:
                                 
                                schedule.append({
                                    "day": self.days[self.days['Day Number'] == day].iloc[0]['Day Name'],
                                    "time": time,
                                    "room": room,
                                    "capacity_room":capacity,
                                    "course": course_info['course'],
                                    "teatcher": course_info['teacher']['name'],
                                    "available":course_info['teacher']['available'],
                                    "day_id":day,
                                    "group":course_info['group'],
                                    "level": course_info['level'],
                                    "dept": course_info['dept'],
                                    "student_count": course_info['std_count']
                                })

            if schedule:
                self.save_to_excel(schedule)
                conflicts = self.check_conflicts(schedule)
            
            # Write conflicts to Excel file
                self.write_conflicts_to_excel(conflicts)
            
                if conflicts:
                    print("\nتم العثور على تعارضات")
                else:
                    print("\n لا يوجد تعارضات")
        else:
            print("❌ لم يتم العثور على حل!")
            print("⚠ حاول تغيير القيود أو عدد القاعات المتاحة.")
            
            # Check for potential conflicts that might be preventing a solution
            initial_conflicts = self.check_initial_conflicts()
            if initial_conflicts:
                print("\nتم العثور على تعارضات محتملة   :")
                self.write_conflicts_to_excel(initial_conflicts)
                print("\n✅ تم حفظ التعارضات المحتملة في ملف 'conflicts.xlsx'")
            else:
                print("\nلم يتم العثور على تعارضات واضحة   ")


if __name__ == "__main__":
    scheduler = TimeTableScheduler()
    scheduler.add_data()
    scheduler.define_variables()
    scheduler.add_constraints()
    scheduler.solve()
